"""
Export Router — Generación de Excel y PDF desde datos de Bind ERP
POST /api/export/excel  → descarga .xlsx
POST /api/export/pdf    → descarga .pdf
"""
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from middleware.auth import get_current_user, CurrentUser

logger = logging.getLogger("atollom.export")

router = APIRouter(prefix="/api/export", tags=["Export"])


# ====================================================================
# SCHEMA DE REQUEST
# ====================================================================
class ExportRequest(BaseModel):
    title: str = "Reporte Bind ERP"
    intent: str = "VENTAS"
    data: List[Dict[str, Any]]
    summary: Optional[str] = None  # Análisis de Gemini incluido en el reporte


# ====================================================================
# HELPER: Aplanar un dict anidado para Excel
# ====================================================================
def _flatten(row: Dict, prefix: str = "") -> Dict:
    flat = {}
    for k, v in row.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            flat.update(_flatten(v, key))
        elif isinstance(v, list):
            flat[key] = str(v)  # Listas como string en Excel
        else:
            flat[key] = v
    return flat


# ====================================================================
# ENDPOINT: Excel
# ====================================================================
@router.post("/excel")
async def export_excel(
    req: ExportRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Genera un archivo .xlsx con los datos y el análisis de Gemini."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado.")

    wb = Workbook()
    ws = wb.active
    ws.title = req.intent[:31]  # Excel limita a 31 chars

    # ── Colores del tema Cybernetic Trust ──
    NAVY   = "001C3E"
    LIME   = "A4DA30"
    WHITE  = "FFFFFF"
    GRAY   = "D0DCE3"

    # ── Fila 1: Título ──
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = req.title
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Fila 2: Metadata ──
    ws.merge_cells("A2:F2")
    meta = ws["A2"]
    meta.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Módulo: {req.intent} | Tenant: {current_user.tenant_id}"
    meta.font = Font(size=9, color=NAVY)
    meta.fill = PatternFill("solid", fgColor=LIME)
    meta.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Resumen de Gemini (si existe) ──
    start_data_row = 4
    if req.summary:
        ws.merge_cells(f"A3:F3")
        summary_cell = ws["A3"]
        summary_cell.value = f"📊 Análisis IA: {req.summary}"
        summary_cell.font = Font(size=9, italic=True, color=NAVY)
        summary_cell.fill = PatternFill("solid", fgColor="E8F5D0")
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[3].height = 45
        start_data_row = 5

    # ── Headers ──
    if not req.data:
        ws[f"A{start_data_row}"] = "No hay datos disponibles."
    else:
        flat_rows = [_flatten(row) for row in req.data]
        headers = list(flat_rows[0].keys())

        header_fill = PatternFill("solid", fgColor=NAVY)
        header_font = Font(bold=True, color=WHITE, size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_data_row, column=col_idx)
            cell.value = header.replace("_", " ").title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # ── Datos ──
        alt_fill = PatternFill("solid", fgColor="F0F5FA")
        for row_idx, row_data in enumerate(flat_rows, start_data_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(key, "")
                cell.border = border
                if fill:
                    cell.fill = fill

        # ── Autoajustar columnas ──
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(header)),
                max((len(str(flat_rows[r].get(header, ""))) for r in range(len(flat_rows))), default=0)
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Guardar en buffer y enviar ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"atollom_{req.intent.lower()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ====================================================================
# ENDPOINT: PDF
# ====================================================================
@router.post("/pdf")
async def export_pdf(
    req: ExportRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Genera un archivo .pdf con análisis y tabla de datos."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no instalado.")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    # ── Colores ──
    NAVY   = colors.HexColor("#001C3E")
    LIME   = colors.HexColor("#A4DA30")
    LIGHT  = colors.HexColor("#E8F5D0")
    GRAY   = colors.HexColor("#D0DCE3")

    styles = getSampleStyleSheet()
    elements = []

    # ── Título ──
    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=NAVY, fontSize=16, spaceAfter=4)
    elements.append(Paragraph(req.title, title_style))

    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], textColor=colors.gray, fontSize=8)
    elements.append(Paragraph(
        f"Módulo: {req.intent} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Tenant: {current_user.tenant_id}",
        meta_style
    ))
    elements.append(Spacer(1, 0.15 * inch))

    # ── Análisis de Gemini ──
    if req.summary:
        insight_style = ParagraphStyle("Insight", parent=styles["Normal"], fontSize=9, leading=13,
                                       backColor=LIGHT, textColor=NAVY, leftIndent=8, rightIndent=8,
                                       spaceBefore=4, spaceAfter=8, borderPadding=6)
        elements.append(Paragraph(f"📊 Análisis IA: {req.summary}", insight_style))
        elements.append(Spacer(1, 0.1 * inch))

    # ── Tabla de datos ──
    if req.data:
        flat_rows = [_flatten(row) for row in req.data]
        headers = list(flat_rows[0].keys())

        table_data = [[h.replace("_", " ").title() for h in headers]]
        for row in flat_rows[:200]:  # Limitar a 200 filas por página
            table_data.append([str(row.get(h, "")) for h in headers])

        col_width = (10 * inch) / max(len(headers), 1)
        col_widths = [min(col_width, 2.5 * inch)] * len(headers)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 8),
            ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
            ("FONTSIZE",   (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FC")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No hay datos disponibles para este reporte.", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    filename = f"atollom_{req.intent.lower()}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
